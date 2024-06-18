import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook

janela = tk.Tk()
janela.resizable(0,0)
janela.geometry('800x500')
janela.title('resGatados')

def opEnt():
    janela2 = tk.Toplevel()
    janela2.resizable(0,0)
    janela2.title('Cadastro de entrada')
    janela2.geometry('500x600')

    dtRsg = tk.Label(
        janela2,
        text = 'Data do resgate',)
    dtRsg.pack()

    dtEntr = tk.Entry(
        janela2,
        width = 15)
    dtEntr.pack()

    resLbl = tk.Label(
        janela2,
        text = 'Origem')
    resLbl.pack()

    resRet = Combobox(janela2)
    resRet['values']= ('Resgate',
                       'Retorno')
    resRet.pack()
    
    cdReg = tk.Label(
        janela2,
        text = 'Código de registro')
    cdReg.pack()

    cdBtn = tk.Entry(
        janela2,
        width = 15)
    cdBtn.pack()

    ttrRsp = tk.Label(
        janela2,
        text = 'Tutor/Responsável')
    ttrRsp.pack()

    ttrEntr = tk.Entry(
        janela2,
        width = 15)
    ttrEntr.pack()

    tlf = tk.Label(
        janela2,
        text = 'Telefone do tutor')
    tlf.pack()

    tlfEntr = tk.Entry(
        janela2,
        width = 15)
    tlfEntr.pack()

    end = tk.Label(
        janela2,
        text = 'Endereço do tutor')
    end.pack()

    endEntr = tk.Entry(
        janela2,
        width = 15)
    endEntr.pack()

    cidd = tk.Label(
        janela2,
        text = 'Cidade do tutor')
    cidd.pack()

    ciddEntr = tk.Entry(
        janela2,
        width = 15)
    ciddEntr.pack()

    orig = tk.Label(
        janela2,
        text = 'Local de origem do gato')
    orig.pack()
    
    origEntr = tk.Entry(
        janela2,
        width = 15)
    origEntr.pack()
    
    corLbl = tk.Label(
        janela2,
        text = 'Cor/Pelagem')
    corLbl.pack()

    cor = Combobox(janela2)
    cor['values']= ('Branco',
                  'Preto',
                  'Frajola',
                  'Siames',
                  'Escaminha',
                  'Tricolor',
                  'Tigrado',
                  'Tigrado com branco',
                  'Cinza',
                  'Laranja',
                  'Cinza com branco',
                  'Laranja com branco',
                  'Outros')
    cor.pack()

    lblCorOlhos = tk.Label(
    janela2,
    text='Cor dos olhos')
    lblCorOlhos.pack()

    corOlho = Combobox(janela2)
    corOlho['values']= ('Amarelo',
                    'Verde',
                    'Azul',
                    'Um de cada cor')
    corOlho.pack()

    lblSex = tk.Label(
    janela2,
    text='Sexo')
    lblSex.pack()

    sexo = Combobox(janela2)
    sexo['values']= ('Macho',
                 'Fêmea',
                 'Não identificado')
    sexo.pack()

    lblId = tk.Label(
    janela2,
    text='Idade')
    lblId.pack()

    id = Combobox(janela2)
    id['values']= ('Bebê',
               'Filhote',
               'Adulto',)
    id.pack()

    lblObs = tk.Label(
    janela2,
    text='Observações')
    lblObs.pack()

    entryObs = tk.Entry(
    janela2,
    width=50)
    entryObs.pack()

    btnCdsi = tk.Button(
    janela2,
    command=cadastroEntrada,
    text = 'Cadastrar')
    btnCdsi.pack()

#def opTr():
#    janela3 = tk.Toplevel()
#    janela3.title('Cadastro de triagem')
#    janela3.geometry('200x200')
#    botao_voltar2 = tk.Button(janela3, text = 'Fechar a janela3', command = janela3.destroy)
#    botao_voltar2.pack()

def opSai():
    janela4 = tk.Toplevel()
    janela4.title('Cadastro de saída')
    janela4.geometry('500x500')
    janela4.resizable(0,0)

    lblIde = tk.Label(
    janela4,
    text='Id da coleira')
    lblIde.pack()

    entryIde = tk.Entry(
    janela4,
    width=50)
    entryIde.pack()

    lblSai = tk.Label(
    janela4,
    text='Motivo da saída')
    lblSai.pack()

    mot = Combobox(janela4)
    mot['values'] = ('Adotado',
                 'Identificado',
                 'Lar temporário',
                 'Outro')
    mot.pack()

    lblObs = tk.Label(
    janela4,
    text='Observações')
    lblObs.pack()

    entryObs = tk.Entry(
    janela4,
    width=50)
    entryObs.pack()

    lblNmTut = tk.Label(
    janela4,
    text='Nome do novo tutor')
    lblNmTut.pack()

    entryNmTut = tk.Entry(
    janela4,
    width=50)
    entryNmTut.pack()

    tutNr = tk.Label(
    janela4,
    text='Número do novo tutor')
    tutNr.pack()

    entryNrTut = tk.Entry(
    janela4,
    width=50)
    entryNrTut.pack()

    lblFanti = tk.Label(
    janela4,
    height=2)
    lblFanti.pack()

    btnCdsi = tk.Button(
    janela4,
    #command=cadastroData,
    text = 'Cadastrar')
    btnCdsi.pack()

def cadastroEntrada():
  try:
        wb = load_workbook('resGatados.xlsx')
  except FileNotFoundError:
        wb = Workbook()

  planilha = wb.active

  getData = dtEntr.get()
  getResRet = resRet.get()
  getCd = cdBtn.get()
  getTtr = ttrEntr.get()
  getTlf = tlfEntr.get()
  getEnd = endEntr.get()
  getCidd = ciddEntr.get()
  getOrig = origEntr.get()
  getCor = cor.get()
  getCorOlho = corOlho.get()
  getSexo = sexo.get()
  getId = id.get()
  getObs = entryObs.get()

  planilha.append(['Data', 'Origem', 'Código de registro', 'Tutor/Responsável', 'Telefone do tutor', 'Endereço do tutor', 'Endereço do tutor', 'Cidade do tutor', 'Local de origem do gato', 'Cor/Pelagem', 'Cor dos olhos', 'Sexo', 'Idade', 'Observações'])
  planilha.append([getData, getResRet, getCd, getTtr, getTlf, getEnd, getCidd, getOrig, getCor, getCorOlho, getSexo, getId, getObs])

  wb.save('resGatados.xlsx')
  
  dtEntr.delete(0, tk.END)
  resRet.delete(0, tk.END)
  cdBtn.delete(0, tk.END)
  ttrEntr.delete(0, tk.END)
  tlfEntr.delete(0, tk.END)
  endEntr.delete(0, tk.END)
  ciddEntr.delete(0, tk.END)
  entryObs.delete(0, tk.END)
    
a = tk.Frame()
b = tk.Frame()
c = tk.Frame()

lblVrd = tk.Label(
    master= a,
    height = 100,
    width = 45,
    text='OnlyCats',
    bg='grey')
lblVrd.pack()

lblSpcs = tk.Label(
    master = c,
    width = 25)
lblSpcs.pack()

btnCds = tk.Button(
    master = b,
    command = opEnt,
    width = 15,
    text = 'Entrada')
btnCds.pack()

#lblSpc = tk.Label(
#    master = b,
#    width = 15)
#lblSpc.pack()

#btnTrg = tk.Button(
#    command = opTr,
#    master = b,
#    width = 15,
#    text = 'Triagem')
#btnTrg.pack()

btnSpc = tk.Label(
    master = b,
    height = 1)
btnSpc.pack()

btnSd = tk.Button(
    command = opSai,
    master = b,
    width = 15,
    text = 'Saida')
btnSd.pack()  

a.pack(side = LEFT)
c.pack(side = LEFT)
b.pack(side = LEFT)

janela.mainloop()
